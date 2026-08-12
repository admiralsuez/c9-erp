"""Excel report generation for the C9 ERP report suite."""

from io import BytesIO
from datetime import datetime
from typing import List, Dict, Optional
import logging

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    pass

logger = logging.getLogger(__name__)


class CustomReportMixin:
    """Custom filtered report generation."""

    def generate_custom_report(self, report_data: Dict) -> bytes:
        """Generate a custom-filtered report Excel workbook."""
        wb = Workbook()
        wb.remove(wb.active)

        period = report_data.get("period", {})

        # --- Orders sheet ---
        orders = report_data.get("orders", [])
        ws_orders = wb.active or wb.create_sheet("Orders")
        ws_orders.title = "Orders"
        headers = ["Order #", "Vendor", "Status", "Created", "Items Count"]
        for col, h in enumerate(headers, 1):
            cell = ws_orders.cell(row=1, column=col, value=h)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = self.border

        for row_idx, o in enumerate(orders, 2):
            ws_orders.cell(row=row_idx, column=1, value=o.get("order_number", "")).border = self.border
            ws_orders.cell(row=row_idx, column=2, value=o.get("vendor_name", "")).border = self.border
            ws_orders.cell(row=row_idx, column=3, value=o.get("status", "")).border = self.border
            ws_orders.cell(row=row_idx, column=4, value=str(o.get("created_at", ""))[:10]).border = self.border
            ws_orders.cell(row=row_idx, column=5, value=o.get("item_count", 0)).border = self.border

            # Items sub-table in notes column
            items = o.get("items", [])
            if items:
                notes = "; ".join(f"{i.get('sku','')} x{i.get('quantity_ordered',0)}" for i in items)
                ws_orders.cell(row=row_idx, column=6, value=notes).border = self.border

        ws_orders.column_dimensions['A'].width = 18
        ws_orders.column_dimensions['B'].width = 22
        ws_orders.column_dimensions['C'].width = 22
        ws_orders.column_dimensions['D'].width = 14
        ws_orders.column_dimensions['E'].width = 12
        ws_orders.column_dimensions['F'].width = 40

        # --- Inventory sheet ---
        inventory = report_data.get("inventory", [])
        ws_inv = wb.create_sheet("Inventory")
        inv_headers = ["SKU", "Name", "Current Qty", "Min Qty", "Reserved", "Category"]
        for col, h in enumerate(inv_headers, 1):
            cell = ws_inv.cell(row=1, column=col, value=h)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = self.border

        for row_idx, i in enumerate(inventory, 2):
            ws_inv.cell(row=row_idx, column=1, value=i.get("sku", "")).border = self.border
            ws_inv.cell(row=row_idx, column=2, value=i.get("name", "")).border = self.border
            ws_inv.cell(row=row_idx, column=3, value=i.get("current_quantity", 0)).border = self.border
            ws_inv.cell(row=row_idx, column=4, value=i.get("minimum_quantity", 0)).border = self.border
            ws_inv.cell(row=row_idx, column=5, value=i.get("reserved_quantity", 0)).border = self.border
            ws_inv.cell(row=row_idx, column=6, value=i.get("category", "")).border = self.border

        ws_inv.column_dimensions['A'].width = 14
        ws_inv.column_dimensions['B'].width = 22
        ws_inv.column_dimensions['C'].width = 12
        ws_inv.column_dimensions['D'].width = 10
        ws_inv.column_dimensions['E'].width = 12
        ws_inv.column_dimensions['F'].width = 18

        # --- Summary sheet ---
        ws_summary = wb.create_sheet("Summary")
        row = 1
        ws_summary.cell(row=row, column=1, value="Custom Report").font = Font(bold=True, size=14)
        row += 1
        ws_summary.cell(row=row, column=1, value=f"Period: {period.get('start','N/A')} to {period.get('end','N/A')}")
        row += 2
        ws_summary.cell(row=row, column=1, value="Metric").font = self.summary_font
        ws_summary.cell(row=row, column=2, value="Value").font = self.summary_font
        row += 1
        for key, val in [("Total Orders", len(orders)), ("Total Items", len(inventory))]:
            ws_summary.cell(row=row, column=1, value=key).border = self.border
            ws_summary.cell(row=row, column=2, value=val).border = self.border
            row += 1

        ws_summary.column_dimensions['A'].width = 25
        ws_summary.column_dimensions['B'].width = 15

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.getvalue()
