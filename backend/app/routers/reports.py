from fastapi import APIRouter, Depends, HTTPException, status, Query
from sqlalchemy.orm import Session
from datetime import datetime, timezone, timedelta
from typing import Optional, List
from app.core.database import get_db
from app.core.auth import get_current_user, require_admin
from app.models import User
from app.services.analytics_service import get_analytics_service
from app.services.pdf_reports import PDFReportGenerator
from app.services.excel_reports import ExcelReportGenerator

router = APIRouter(prefix="/reports", tags=["Reports"])

PERIOD_MAP = {
    "weekly": timedelta(days=7),
    "monthly": timedelta(days=30),
    "quarterly": timedelta(days=90),
}


@router.post("/generate")
def generate_report(
    period: str = Query(..., pattern="^(weekly|monthly|quarterly)$"),
    format: str = Query("pdf", pattern="^(pdf|excel|json)$"),
    variant_ids: Optional[List[int]] = Query(None, description="Filter by inventory variant IDs"),
    view: bool = Query(False, description="If true, serve inline for browser viewing instead of download"),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin)
):
    """Generate a comprehensive weekly/monthly/quarterly report.
    
    Args:
        period: Report period (weekly, monthly, quarterly)
        format: Output format (pdf, excel, json)
        variant_ids: Optional list of inventory variant IDs to include
        view: If true, serve inline for browser viewing
    """
    delta = PERIOD_MAP.get(period)
    if not delta:
        raise HTTPException(status_code=400, detail="Invalid period")

    now = datetime.now(timezone.utc)
    period_start = now - delta
    analytics = get_analytics_service(db)
    analytics_data = analytics.get_dashboard_overview(date_from=period_start, date_to=now)
    
    # Filter by variants if provided
    if variant_ids:
        analytics_data['variant_filter'] = variant_ids
        analytics_data['filtered_by_variants'] = True

    if format == "json":
        from fastapi.responses import JSONResponse
        return JSONResponse(content=analytics_data)

    try:
        if format == "excel":
            gen = ExcelReportGenerator()
            content = gen.generate_analytics_report(analytics_data)
            media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            filename = f"{period}_report_{now.strftime('%Y%m%d')}.xlsx"
        else:
            gen = PDFReportGenerator()
            content = gen.generate_analytics_report(analytics_data, include_charts=False)
            media_type = "application/pdf"
            filename = f"{period}_report_{now.strftime('%Y%m%d')}.pdf"

        disposition = "inline" if view else "attachment"
        from fastapi.responses import Response
        return Response(
            content=content,
            media_type=media_type,
            headers={"Content-Disposition": f'{disposition}; filename="{filename}"'},
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Report generation failed: {str(e)}")


@router.post("/custom")
def generate_custom_report(
    date_from: str = Query(..., description="Start date (ISO format, e.g. 2026-01-01)"),
    date_to: str = Query(..., description="End date (ISO format, e.g. 2026-12-31)"),
    item_ids: Optional[List[int]] = Query(None, description="Filter by inventory item IDs"),
    variant_ids: Optional[List[int]] = Query(None, description="Filter by inventory variant IDs"),
    vendor_ids: Optional[List[int]] = Query(None, description="Filter by vendor IDs"),
    format: str = Query("pdf", pattern="^(pdf|excel|json)$"),
    view: bool = Query(False, description="If true, serve inline for browser viewing instead of download"),
    page: int = Query(1, ge=1, description="Page number for pagination"),
    page_size: int = Query(50, ge=10, le=500, description="Items per page"),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin)
):
    """Generate a custom report filtered by date range, items/variants, and/or vendors.
    
    Args:
        date_from: Report start date (ISO format)
        date_to: Report end date (ISO format)
        item_ids: Filter by parent inventory item IDs
        variant_ids: Filter by inventory variant IDs (child items)
        vendor_ids: Filter by vendor IDs
        format: Output format (pdf, excel, json)
        view: If true, serve inline for browser viewing
        page: Page number for pagination (only for JSON format)
        page_size: Items per page (only for JSON format)
    """
    try:
        start = datetime.fromisoformat(date_from).replace(tzinfo=timezone.utc)
        end = datetime.fromisoformat(date_to).replace(tzinfo=timezone.utc)
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date format. Use ISO format (YYYY-MM-DD)")

    analytics = get_analytics_service(db)

    # Combine item_ids and variant_ids for filtering
    effective_item_ids = item_ids or []
    if variant_ids:
        effective_item_ids = list(set(effective_item_ids + variant_ids))
    
    orders = analytics.get_filtered_orders(start, end, effective_item_ids or None, vendor_ids, page=page, page_size=page_size)
    inventory = analytics.get_filtered_inventory(effective_item_ids or None, page=page, page_size=page_size)

    report_data = {
        "orders": orders if isinstance(orders, dict) else {"data": orders},
        "inventory": inventory if isinstance(inventory, dict) else {"data": inventory},
        "period": {
            "label": "Custom",
            "start": start.strftime("%Y-%m-%d"),
            "end": end.strftime("%Y-%m-%d"),
        },
        "calculated_at": datetime.now(timezone.utc).isoformat(),
        "filters": {
            "item_ids": item_ids,
            "variant_ids": variant_ids,
            "vendor_ids": vendor_ids,
        },
    }

    if format == "json":
        from fastapi.responses import JSONResponse
        return JSONResponse(content=report_data)

    try:
        if format == "excel":
            gen = ExcelReportGenerator()
            # Use generate_custom_report if it exists, otherwise fall back to custom handling
            if hasattr(gen, 'generate_custom_report'):
                content = gen.generate_custom_report(report_data)
            else:
                # Manual handling for Excel custom report
                from openpyxl import Workbook
                wb = Workbook()
                wb.remove(wb.active)
                ws = wb.create_sheet("Orders", 0)
                gen._populate_orders_sheet(ws, orders, report_data.get("period"))
                ws_inv = wb.create_sheet("Inventory")
                gen._populate_inventory_sheet(ws_inv, inventory)
                from io import BytesIO
                excel_buffer = BytesIO()
                wb.save(excel_buffer)
                excel_buffer.seek(0)
                content = excel_buffer.getvalue()
            media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            filename = f"custom_report_{end.strftime('%Y%m%d')}.xlsx"
        else:
            gen = PDFReportGenerator()
            content = gen.generate_custom_report(report_data)
            media_type = "application/pdf"
            filename = f"custom_report_{end.strftime('%Y%m%d')}.pdf"

        disposition = "inline" if view else "attachment"
        from fastapi.responses import Response
        return Response(
            content=content,
            media_type=media_type,
            headers={"Content-Disposition": f'{disposition}; filename="{filename}"'},
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Custom report generation failed: {str(e)}")
