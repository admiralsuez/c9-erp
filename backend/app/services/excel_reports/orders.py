"""Excel report generation for the C9 ERP report suite."""

from io import BytesIO
from datetime import datetime
from typing import List, Dict, Optional
import logging

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    pass

logger = logging.getLogger(__name__)


class OrderReportMixin:
    """Order report generation."""

    def generate_order_report(
        self,
        orders_data: List[Dict],
        date_range: Optional[Dict] = None,
        summary_stats: Optional[Dict] = None
    ) -> bytes:
        """
        Generate orders report as Excel workbook.
        
        Args:
            orders_data: List of order dictionaries
            date_range: Optional dict with 'start' and 'end' dates
            summary_stats: Optional summary statistics dict
            
        Returns:
            Excel file as bytes
        """
        wb = Workbook()
        wb.remove(wb.active)
        
        # Summary sheet
        if summary_stats:
            self._create_summary_sheet(wb, summary_stats, "Summary")
        
        # Orders sheet
        ws_orders = wb.create_sheet("Orders", 0)
        self._populate_orders_sheet(ws_orders, orders_data, date_range)
        
        # Return bytes
        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        return excel_buffer.getvalue()

    def _populate_orders_sheet(self, ws, orders_data: List[Dict], date_range: Optional[Dict] = None):
        """Populate orders worksheet with creator, approver, items as alternating columns."""
        # Header with date range
        if date_range:
            ws['A1'] = f"Order Report: {date_range.get('start')} to {date_range.get('end')}"
            ws['A1'].font = Font(bold=True, size=12)
            ws.merge_cells('A1:H1')
            start_row = 3
        else:
            start_row = 1
        
        # Column headers: Date | Vendor | Status | Created By | Approved By | Items... | Total Count
        headers = ["Date", "Vendor", "Status", "Created By", "Approved By", "Items", "Qty", "Item Count"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=start_row, column=col)
            cell.value = header
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.border = self.border
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Data rows
        row = start_row + 1
        for order in orders_data:
            created_date = str(order.get('created_at', ''))[:10]
            ws.cell(row=row, column=1).value = created_date
            ws.cell(row=row, column=2).value = order.get('vendor_name', '')
            ws.cell(row=row, column=3).value = order.get('status', '')
            ws.cell(row=row, column=4).value = order.get('created_by', '')
            ws.cell(row=row, column=5).value = order.get('approved_by', '')
            
            # Get items and quantities
            items = order.get('items', [])
            total_qty = 0
            items_str = ""
            
            if items:
                item_names = []
                for item in items:
                    item_name = item.get('name', f"Item #{item.get('id', '')}")
                    qty = item.get('quantity_ordered', 0)
                    total_qty += qty
                    item_names.append(item_name)
                items_str = " | ".join(item_names)
            
            ws.cell(row=row, column=6).value = items_str
            ws.cell(row=row, column=7).value = total_qty if total_qty > 0 else order.get('item_count', 0)
            ws.cell(row=row, column=8).value = len(items) if items else order.get('item_count', 0)
            
            for col in range(1, 9):
                cell = ws.cell(row=row, column=col)
                cell.border = self.border
                if col in [1, 7, 8]:  # Right-align dates and numbers
                    cell.alignment = Alignment(horizontal="right")
                else:
                    cell.alignment = Alignment(horizontal="left")
            
            row += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 12  # Date
        ws.column_dimensions['B'].width = 20  # Vendor
        ws.column_dimensions['C'].width = 12  # Status
        ws.column_dimensions['D'].width = 15  # Created By
        ws.column_dimensions['E'].width = 15  # Approved By
        ws.column_dimensions['F'].width = 35  # Items
        ws.column_dimensions['G'].width = 8   # Qty
        ws.column_dimensions['H'].width = 12  # Item Count
