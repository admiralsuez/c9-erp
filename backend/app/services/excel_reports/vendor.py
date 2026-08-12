"""Excel report generation for the C9 ERP report suite."""

from io import BytesIO
from datetime import datetime
from typing import List, Dict, Optional
import logging

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    pass

logger = logging.getLogger(__name__)


class VendorReportMixin:
    """Vendor performance report generation."""

    def generate_vendor_report(
        self,
        vendor_data: List[Dict],
        performance_metrics: Optional[Dict] = None
    ) -> bytes:
        """
        Generate vendor performance report as Excel workbook.
        
        Args:
            vendor_data: List of vendor dictionaries
            performance_metrics: Optional performance metrics
            
        Returns:
            Excel file as bytes
        """
        wb = Workbook()
        wb.remove(wb.active)
        
        # Summary sheet
        if performance_metrics:
            self._create_summary_sheet(wb, performance_metrics, "Summary")
        
        # Vendors sheet
        ws_vendors = wb.create_sheet("Vendors", 0)
        self._populate_vendors_sheet(ws_vendors, vendor_data)
        
        # Return bytes
        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        return excel_buffer.getvalue()

    def _populate_vendors_sheet(self, ws, vendor_data: List[Dict]):
        """Populate vendors worksheet."""
        # Column headers
        headers = ["Vendor Name", "Order Count", "On-Time %", "Delivery Performance"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.border = self.border
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Data rows
        row = 2
        for vendor in vendor_data:
            ws.cell(row=row, column=1).value = vendor.get('vendor_name', '')
            ws.cell(row=row, column=2).value = vendor.get('order_count', 0)
            on_time_pct = vendor.get('on_time_percentage', 0)
            ws.cell(row=row, column=3).value = on_time_pct / 100  # Format as percentage
            ws.cell(row=row, column=4).value = "Good" if on_time_pct >= 90 else "Fair" if on_time_pct >= 75 else "Poor"
            
            for col in range(1, 5):
                cell = ws.cell(row=row, column=col)
                cell.border = self.border
                if col == 3:
                    cell.number_format = '0%'
                    cell.alignment = Alignment(horizontal="right")
                elif col in [2]:
                    cell.alignment = Alignment(horizontal="right")
            
            row += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 14
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 20
