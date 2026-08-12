"""Excel report generation for the C9 ERP report suite."""

from io import BytesIO
from datetime import datetime
from typing import List, Dict, Optional
import logging

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    pass

logger = logging.getLogger(__name__)


class AnalyticsReportMixin:
    """Analytics report generation."""

    def generate_analytics_report(
        self,
        analytics_data: Dict
    ) -> bytes:
        """
        Generate comprehensive analytics report as Excel workbook.
        
        Args:
            analytics_data: Complete analytics dictionary
            
        Returns:
            Excel file as bytes
        """
        wb = Workbook()
        wb.remove(wb.active)
        
        # Overview sheet
        ws_overview = wb.create_sheet("Overview", 0)
        self._create_analytics_overview_sheet(ws_overview, analytics_data)
        
        # Order metrics sheet
        if 'order_metrics' in analytics_data:
            ws_orders = wb.create_sheet("Order Metrics")
            order_metrics = analytics_data['order_metrics']
            summary = {
                "Total Orders": order_metrics.get('total_orders', 0),
                "By Status": str(order_metrics.get('by_status', {})),
                "Pending Approvals": order_metrics.get('pending_approvals', 0),
                "Avg Approval Time (days)": order_metrics.get('average_approval_time_days', 0),
                "Avg Dispatch Time (days)": order_metrics.get('average_dispatch_time_days', 0),
            }
            self._populate_summary_sheet(ws_orders, summary)
        
        # Inventory sheet
        if 'inventory_health' in analytics_data:
            ws_inventory = wb.create_sheet("Inventory")
            inv_health = analytics_data['inventory_health']
            summary = {
                "Total Items": inv_health.get('total_items', 0),
                "Low Stock Count": inv_health.get('low_stock_count', 0),
                "Total Quantity": inv_health.get('total_quantity', 0),
            }
            self._populate_summary_sheet(ws_inventory, summary)
            
            if inv_health.get('low_stock_items'):
                self._populate_low_stock_sheet(
                    ws_inventory,
                    inv_health['low_stock_items'],
                    start_row=6
                )
        
        # Vendor sheet
        if 'vendor_performance' in analytics_data:
            ws_vendors = wb.create_sheet("Vendor Performance")
            self._populate_vendors_sheet(ws_vendors, analytics_data['vendor_performance'])
        
        # Email stats sheet
        if 'email_stats' in analytics_data:
            ws_emails = wb.create_sheet("Email Statistics")
            email_stats = analytics_data['email_stats']
            summary = {
                "Total Emails": email_stats.get('total_emails', 0),
                "Sent": email_stats.get('sent_count', 0),
                "Failed": email_stats.get('failed_count', 0),
                "Period (days)": email_stats.get('period_days', 0),
            }
            self._populate_summary_sheet(ws_emails, summary)
        
        # User activity sheet
        if 'user_activity' in analytics_data:
            ws_users = wb.create_sheet("User Activity")
            user_activity = analytics_data['user_activity']
            summary = {
                "Active Users": user_activity.get('active_users', 0),
                "Total Actions": user_activity.get('total_actions', 0),
                "Orders Created": user_activity.get('orders_created', 0),
                "Period (days)": user_activity.get('period_days', 0),
            }
            self._populate_summary_sheet(ws_users, summary)
        
        # Return bytes
        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        return excel_buffer.getvalue()

    def _create_analytics_overview_sheet(self, ws, analytics_data: Dict):
        """Create overview sheet for analytics report."""
        ws['A1'] = f"{self.company_name} Analytics Report"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:D1')
        
        ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws['A2'].font = Font(italic=True, size=10)
        
        row = 4
        
        # Order Metrics
        if 'order_metrics' in analytics_data:
            ws[f'A{row}'] = "Order Metrics"
            ws[f'A{row}'].font = Font(bold=True, size=11)
            row += 1
            
            order_stats = analytics_data['order_metrics']
            data = {
                "Total Orders": order_stats.get('total_orders', 0),
                "Pending Approvals": order_stats.get('pending_approvals', 0),
                "Avg Approval Time (days)": round(order_stats.get('average_approval_time_days', 0), 2),
                "Avg Dispatch Time (days)": round(order_stats.get('average_dispatch_time_days', 0), 2),
            }
            
            for key, value in data.items():
                ws[f'A{row}'] = key
                ws[f'B{row}'] = value
                row += 1
            
            row += 1
        
        # Inventory Health
        if 'inventory_health' in analytics_data:
            ws[f'A{row}'] = "Inventory Health"
            ws[f'A{row}'].font = Font(bold=True, size=11)
            row += 1
            
            inv_health = analytics_data['inventory_health']
            data = {
                "Total Items": inv_health.get('total_items', 0),
                "Low Stock Items": inv_health.get('low_stock_count', 0),
                "Total Quantity": round(float(inv_health.get('total_quantity', 0)), 0),
            }
            
            for key, value in data.items():
                ws[f'A{row}'] = key
                ws[f'B{row}'] = value
                row += 1
            
            row += 1
        
        # Email Stats
        if 'email_stats' in analytics_data:
            ws[f'A{row}'] = "Email Statistics"
            ws[f'A{row}'].font = Font(bold=True, size=11)
            row += 1
            
            email_stats = analytics_data['email_stats']
            data = {
                "Total Emails": email_stats.get('total_emails', 0),
                "Sent": email_stats.get('sent_count', 0),
                "Failed": email_stats.get('failed_count', 0),
            }
            
            for key, value in data.items():
                ws[f'A{row}'] = key
                ws[f'B{row}'] = value
                row += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 15
