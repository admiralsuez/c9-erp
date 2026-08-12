"""Excel report generation for the C9 ERP report suite."""

from io import BytesIO
from datetime import datetime
from typing import List, Dict, Optional
import logging

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    pass

logger = logging.getLogger(__name__)


class ExcelReportBaseMixin:
    """Shared state, styles and helpers for Excel report generation."""

    def __init__(self, company_name: str = "Cloud9 ERP"):
        self.company_name = company_name
        self.header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        self.header_font = Font(bold=True, color="FFFFFF", size=11)
        self.summary_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        self.summary_font = Font(bold=True, size=10)
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    def _create_summary_sheet(self, wb, summary_dict: Dict, sheet_name: str = "Summary"):
        """Create a summary statistics sheet."""
        ws = wb.create_sheet(sheet_name, 0)
        self._populate_summary_sheet(ws, summary_dict)

    def _populate_summary_sheet(self, ws, summary_dict: Dict):
        """Populate a worksheet with summary data."""
        row = 1
        ws[f'A{row}'] = "Metric"
        ws[f'B{row}'] = "Value"
        
        # Format header
        for col in ['A', 'B']:
            cell = ws[f'{col}{row}']
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.border = self.border
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Data rows
        row = 2
        for key, value in summary_dict.items():
            ws[f'A{row}'] = key
            if isinstance(value, (int, float)):
                ws[f'B{row}'] = float(value) if isinstance(value, float) else int(value)
            else:
                ws[f'B{row}'] = str(value)
            
            for col in ['A', 'B']:
                cell = ws[f'{col}{row}']
                cell.border = self.border
                cell.alignment = Alignment(horizontal="left" if col == 'A' else "right")
            
            row += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20

    def _populate_low_stock_sheet(self, ws, items: List[Dict], start_row: int = 1):
        """Populate low stock items worksheet."""
        # Column headers
        headers = ["SKU", "Item Name", "Current Qty", "Minimum Qty", "Variance"]
        row = start_row
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
            cell.border = self.border
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Data rows
        row = start_row + 1
        for item in items:
            current = float(item.get('current', 0))
            minimum = float(item.get('minimum', 0))
            variance = current - minimum
            
            ws.cell(row=row, column=1).value = item.get('sku', '')
            ws.cell(row=row, column=2).value = item.get('name', '')
            ws.cell(row=row, column=3).value = current
            ws.cell(row=row, column=4).value = minimum
            ws.cell(row=row, column=5).value = variance
            
            for col in range(1, 6):
                cell = ws.cell(row=row, column=col)
                cell.border = self.border
                if col in [3, 4, 5]:
                    cell.alignment = Alignment(horizontal="right")
                if col == 5 and variance < 0:
                    cell.fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
            
            row += 1
        
        # Adjust column widths
        if start_row == 1:
            ws.column_dimensions['A'].width = 12
            ws.column_dimensions['B'].width = 25
            ws.column_dimensions['C'].width = 14
            ws.column_dimensions['D'].width = 14
            ws.column_dimensions['E'].width = 12
