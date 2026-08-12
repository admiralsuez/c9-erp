"""Excel report generation for the C9 ERP report suite."""

from io import BytesIO
from datetime import datetime
from typing import List, Dict, Optional
import logging

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    pass

logger = logging.getLogger(__name__)


class InventoryReportMixin:
    """Inventory report generation."""

    def generate_inventory_report(
        self,
        inventory_data: List[Dict],
        summary_stats: Optional[Dict] = None,
        low_stock_items: Optional[List[Dict]] = None
    ) -> bytes:
        """
        Generate inventory report as Excel workbook.
        
        Args:
            inventory_data: List of inventory item dictionaries
            summary_stats: Optional summary statistics
            low_stock_items: Optional list of low stock items
            
        Returns:
            Excel file as bytes
        """
        wb = Workbook()
        wb.remove(wb.active)
        
        # Summary sheet
        if summary_stats:
            self._create_summary_sheet(wb, summary_stats, "Summary")
        
        # All inventory sheet
        ws_inventory = wb.create_sheet("All Inventory", 0)
        self._populate_inventory_sheet(ws_inventory, inventory_data)
        
        # Low stock sheet
        if low_stock_items:
            ws_low_stock = wb.create_sheet("Low Stock")
            self._populate_low_stock_sheet(ws_low_stock, low_stock_items)
        
        # Return bytes
        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        return excel_buffer.getvalue()

    def _populate_inventory_sheet(self, ws, inventory_data: List[Dict]):
        """Populate inventory worksheet with opening, closing, sent, restocked."""
        # Column headers: SKU | Name | Opening | Received/Restocked | Sent/Dispatched | Closing | Status
        headers = ["SKU", "Name", "Opening Qty", "Restocked", "Sent", "Closing Qty", "Min Qty", "Status"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.border = self.border
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Data rows
        row = 2
        for item in inventory_data:
            current = float(item.get('current_quantity', 0))
            minimum = float(item.get('minimum_quantity', 0))
            reserved = float(item.get('reserved_quantity', 0))
            
            # Calculate opening as current (in production, this would come from ledger start date)
            opening_qty = float(item.get('opening_quantity', current))
            # Restocked = items added since opening
            restocked_qty = float(item.get('restocked_quantity', 0))
            # Sent/Dispatched = items removed since opening
            sent_qty = float(item.get('sent_quantity', 0))
            # Closing = current quantity
            closing_qty = current
            
            # Status: OK if current > minimum, Low otherwise
            status = "Low" if closing_qty <= minimum else "OK"
            
            ws.cell(row=row, column=1).value = item.get('sku', '')
            ws.cell(row=row, column=2).value = item.get('name', '')
            ws.cell(row=row, column=3).value = opening_qty
            ws.cell(row=row, column=4).value = restocked_qty
            ws.cell(row=row, column=5).value = sent_qty
            ws.cell(row=row, column=6).value = closing_qty
            ws.cell(row=row, column=7).value = minimum
            ws.cell(row=row, column=8).value = status
            
            for col in range(1, 9):
                cell = ws.cell(row=row, column=col)
                cell.border = self.border
                if col in [3, 4, 5, 6, 7]:  # Right-align numbers
                    cell.alignment = Alignment(horizontal="right")
                elif col == 8:
                    # Color code status
                    if status == "Low":
                        cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                    else:
                        cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            
            row += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 12  # SKU
        ws.column_dimensions['B'].width = 25  # Name
        ws.column_dimensions['C'].width = 14  # Opening Qty
        ws.column_dimensions['D'].width = 14  # Restocked
        ws.column_dimensions['E'].width = 12  # Sent
        ws.column_dimensions['F'].width = 14  # Closing Qty
        ws.column_dimensions['G'].width = 10  # Min Qty
        ws.column_dimensions['H'].width = 10  # Status
