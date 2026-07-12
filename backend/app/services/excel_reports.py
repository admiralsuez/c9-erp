"""
Excel Report Generation Service for Phase 7.

Generates professional Excel reports with multiple sheets, formatting, and formulas.
"""

from io import BytesIO
from datetime import datetime
from typing import List, Dict, Optional
import logging

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    pass

logger = logging.getLogger(__name__)


class ExcelReportGenerator:
    """Generate professional Excel reports with formatting and charts."""
    
    def __init__(self, company_name: str = "Cloud9 ERP"):
        self.company_name = company_name
        self.header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        self.header_font = Font(bold=True, color="FFFFFF", size=11)
        self.summary_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        self.summary_font = Font(bold=True, size=10)
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def generate_order_report(
        self,
        orders_data: List[Dict],
        date_range: Optional[Dict] = None,
        summary_stats: Optional[Dict] = None
    ) -> bytes:
        """
        Generate orders report as Excel workbook.
        
        Args:
            orders_data: List of order dictionaries
            date_range: Optional dict with 'start' and 'end' dates
            summary_stats: Optional summary statistics dict
            
        Returns:
            Excel file as bytes
        """
        wb = Workbook()
        wb.remove(wb.active)
        
        # Summary sheet
        if summary_stats:
            self._create_summary_sheet(wb, summary_stats, "Summary")
        
        # Orders sheet
        ws_orders = wb.create_sheet("Orders", 0)
        self._populate_orders_sheet(ws_orders, orders_data, date_range)
        
        # Return bytes
        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        return excel_buffer.getvalue()
    
    def generate_inventory_report(
        self,
        inventory_data: List[Dict],
        summary_stats: Optional[Dict] = None,
        low_stock_items: Optional[List[Dict]] = None
    ) -> bytes:
        """
        Generate inventory report as Excel workbook.
        
        Args:
            inventory_data: List of inventory item dictionaries
            summary_stats: Optional summary statistics
            low_stock_items: Optional list of low stock items
            
        Returns:
            Excel file as bytes
        """
        wb = Workbook()
        wb.remove(wb.active)
        
        # Summary sheet
        if summary_stats:
            self._create_summary_sheet(wb, summary_stats, "Summary")
        
        # All inventory sheet
        ws_inventory = wb.create_sheet("All Inventory", 0)
        self._populate_inventory_sheet(ws_inventory, inventory_data)
        
        # Low stock sheet
        if low_stock_items:
            ws_low_stock = wb.create_sheet("Low Stock")
            self._populate_low_stock_sheet(ws_low_stock, low_stock_items)
        
        # Return bytes
        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        return excel_buffer.getvalue()
    
    def generate_vendor_report(
        self,
        vendor_data: List[Dict],
        performance_metrics: Optional[Dict] = None
    ) -> bytes:
        """
        Generate vendor performance report as Excel workbook.
        
        Args:
            vendor_data: List of vendor dictionaries
            performance_metrics: Optional performance metrics
            
        Returns:
            Excel file as bytes
        """
        wb = Workbook()
        wb.remove(wb.active)
        
        # Summary sheet
        if performance_metrics:
            self._create_summary_sheet(wb, performance_metrics, "Summary")
        
        # Vendors sheet
        ws_vendors = wb.create_sheet("Vendors", 0)
        self._populate_vendors_sheet(ws_vendors, vendor_data)
        
        # Return bytes
        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        return excel_buffer.getvalue()
    
    def generate_analytics_report(
        self,
        analytics_data: Dict
    ) -> bytes:
        """
        Generate comprehensive analytics report as Excel workbook.
        
        Args:
            analytics_data: Complete analytics dictionary
            
        Returns:
            Excel file as bytes
        """
        wb = Workbook()
        wb.remove(wb.active)
        
        # Overview sheet
        ws_overview = wb.create_sheet("Overview", 0)
        self._create_analytics_overview_sheet(ws_overview, analytics_data)
        
        # Order metrics sheet
        if 'order_metrics' in analytics_data:
            ws_orders = wb.create_sheet("Order Metrics")
            order_metrics = analytics_data['order_metrics']
            summary = {
                "Total Orders": order_metrics.get('total_orders', 0),
                "By Status": str(order_metrics.get('by_status', {})),
                "Pending Approvals": order_metrics.get('pending_approvals', 0),
                "Avg Approval Time (days)": order_metrics.get('average_approval_time_days', 0),
                "Avg Dispatch Time (days)": order_metrics.get('average_dispatch_time_days', 0),
            }
            self._populate_summary_sheet(ws_orders, summary)
        
        # Inventory sheet
        if 'inventory_health' in analytics_data:
            ws_inventory = wb.create_sheet("Inventory")
            inv_health = analytics_data['inventory_health']
            summary = {
                "Total Items": inv_health.get('total_items', 0),
                "Low Stock Count": inv_health.get('low_stock_count', 0),
                "Total Quantity": inv_health.get('total_quantity', 0),
            }
            self._populate_summary_sheet(ws_inventory, summary)
            
            if inv_health.get('low_stock_items'):
                self._populate_low_stock_sheet(
                    ws_inventory,
                    inv_health['low_stock_items'],
                    start_row=6
                )
        
        # Vendor sheet
        if 'vendor_performance' in analytics_data:
            ws_vendors = wb.create_sheet("Vendor Performance")
            self._populate_vendors_sheet(ws_vendors, analytics_data['vendor_performance'])
        
        # Email stats sheet
        if 'email_stats' in analytics_data:
            ws_emails = wb.create_sheet("Email Statistics")
            email_stats = analytics_data['email_stats']
            summary = {
                "Total Emails": email_stats.get('total_emails', 0),
                "Sent": email_stats.get('sent_count', 0),
                "Failed": email_stats.get('failed_count', 0),
                "Period (days)": email_stats.get('period_days', 0),
            }
            self._populate_summary_sheet(ws_emails, summary)
        
        # User activity sheet
        if 'user_activity' in analytics_data:
            ws_users = wb.create_sheet("User Activity")
            user_activity = analytics_data['user_activity']
            summary = {
                "Active Users": user_activity.get('active_users', 0),
                "Total Actions": user_activity.get('total_actions', 0),
                "Orders Created": user_activity.get('orders_created', 0),
                "Period (days)": user_activity.get('period_days', 0),
            }
            self._populate_summary_sheet(ws_users, summary)
        
        # Return bytes
        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        return excel_buffer.getvalue()
    
    # ============ SHEET BUILDERS ============
    
    def _create_summary_sheet(self, wb, summary_dict: Dict, sheet_name: str = "Summary"):
        """Create a summary statistics sheet."""
        ws = wb.create_sheet(sheet_name, 0)
        self._populate_summary_sheet(ws, summary_dict)
    
    def _populate_summary_sheet(self, ws, summary_dict: Dict):
        """Populate a worksheet with summary data."""
        row = 1
        ws[f'A{row}'] = "Metric"
        ws[f'B{row}'] = "Value"
        
        # Format header
        for col in ['A', 'B']:
            cell = ws[f'{col}{row}']
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.border = self.border
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Data rows
        row = 2
        for key, value in summary_dict.items():
            ws[f'A{row}'] = key
            if isinstance(value, (int, float)):
                ws[f'B{row}'] = float(value) if isinstance(value, float) else int(value)
            else:
                ws[f'B{row}'] = str(value)
            
            for col in ['A', 'B']:
                cell = ws[f'{col}{row}']
                cell.border = self.border
                cell.alignment = Alignment(horizontal="left" if col == 'A' else "right")
            
            row += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20
    
    def _populate_orders_sheet(self, ws, orders_data: List[Dict], date_range: Optional[Dict] = None):
        """Populate orders worksheet with creator, approver, items as alternating columns."""
        # Header with date range
        if date_range:
            ws['A1'] = f"Order Report: {date_range.get('start')} to {date_range.get('end')}"
            ws['A1'].font = Font(bold=True, size=12)
            ws.merge_cells('A1:H1')
            start_row = 3
        else:
            start_row = 1
        
        # Column headers: Date | Vendor | Status | Created By | Approved By | Items... | Total Count
        headers = ["Date", "Vendor", "Status", "Created By", "Approved By", "Items", "Qty", "Item Count"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=start_row, column=col)
            cell.value = header
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.border = self.border
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Data rows
        row = start_row + 1
        for order in orders_data:
            created_date = str(order.get('created_at', ''))[:10]
            ws.cell(row=row, column=1).value = created_date
            ws.cell(row=row, column=2).value = order.get('vendor_name', '')
            ws.cell(row=row, column=3).value = order.get('status', '')
            ws.cell(row=row, column=4).value = order.get('created_by', '')
            ws.cell(row=row, column=5).value = order.get('approved_by', '')
            
            # Get items and quantities
            items = order.get('items', [])
            total_qty = 0
            items_str = ""
            
            if items:
                item_names = []
                for item in items:
                    item_name = item.get('name', f"Item #{item.get('id', '')}")
                    qty = item.get('quantity_ordered', 0)
                    total_qty += qty
                    item_names.append(item_name)
                items_str = " | ".join(item_names)
            
            ws.cell(row=row, column=6).value = items_str
            ws.cell(row=row, column=7).value = total_qty if total_qty > 0 else order.get('item_count', 0)
            ws.cell(row=row, column=8).value = len(items) if items else order.get('item_count', 0)
            
            for col in range(1, 9):
                cell = ws.cell(row=row, column=col)
                cell.border = self.border
                if col in [1, 7, 8]:  # Right-align dates and numbers
                    cell.alignment = Alignment(horizontal="right")
                else:
                    cell.alignment = Alignment(horizontal="left")
            
            row += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 12  # Date
        ws.column_dimensions['B'].width = 20  # Vendor
        ws.column_dimensions['C'].width = 12  # Status
        ws.column_dimensions['D'].width = 15  # Created By
        ws.column_dimensions['E'].width = 15  # Approved By
        ws.column_dimensions['F'].width = 35  # Items
        ws.column_dimensions['G'].width = 8   # Qty
        ws.column_dimensions['H'].width = 12  # Item Count
    
    def _populate_inventory_sheet(self, ws, inventory_data: List[Dict]):
        """Populate inventory worksheet with opening, closing, sent, restocked."""
        # Column headers: SKU | Name | Opening | Received/Restocked | Sent/Dispatched | Closing | Status
        headers = ["SKU", "Name", "Opening Qty", "Restocked", "Sent", "Closing Qty", "Min Qty", "Status"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.border = self.border
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Data rows
        row = 2
        for item in inventory_data:
            current = float(item.get('current_quantity', 0))
            minimum = float(item.get('minimum_quantity', 0))
            reserved = float(item.get('reserved_quantity', 0))
            
            # Calculate opening as current (in production, this would come from ledger start date)
            opening_qty = float(item.get('opening_quantity', current))
            # Restocked = items added since opening
            restocked_qty = float(item.get('restocked_quantity', 0))
            # Sent/Dispatched = items removed since opening
            sent_qty = float(item.get('sent_quantity', 0))
            # Closing = current quantity
            closing_qty = current
            
            # Status: OK if current > minimum, Low otherwise
            status = "Low" if closing_qty <= minimum else "OK"
            
            ws.cell(row=row, column=1).value = item.get('sku', '')
            ws.cell(row=row, column=2).value = item.get('name', '')
            ws.cell(row=row, column=3).value = opening_qty
            ws.cell(row=row, column=4).value = restocked_qty
            ws.cell(row=row, column=5).value = sent_qty
            ws.cell(row=row, column=6).value = closing_qty
            ws.cell(row=row, column=7).value = minimum
            ws.cell(row=row, column=8).value = status
            
            for col in range(1, 9):
                cell = ws.cell(row=row, column=col)
                cell.border = self.border
                if col in [3, 4, 5, 6, 7]:  # Right-align numbers
                    cell.alignment = Alignment(horizontal="right")
                elif col == 8:
                    # Color code status
                    if status == "Low":
                        cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                    else:
                        cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            
            row += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 12  # SKU
        ws.column_dimensions['B'].width = 25  # Name
        ws.column_dimensions['C'].width = 14  # Opening Qty
        ws.column_dimensions['D'].width = 14  # Restocked
        ws.column_dimensions['E'].width = 12  # Sent
        ws.column_dimensions['F'].width = 14  # Closing Qty
        ws.column_dimensions['G'].width = 10  # Min Qty
        ws.column_dimensions['H'].width = 10  # Status
    
    def _populate_vendors_sheet(self, ws, vendor_data: List[Dict]):
        """Populate vendors worksheet."""
        # Column headers
        headers = ["Vendor Name", "Order Count", "On-Time %", "Delivery Performance"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.border = self.border
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Data rows
        row = 2
        for vendor in vendor_data:
            ws.cell(row=row, column=1).value = vendor.get('vendor_name', '')
            ws.cell(row=row, column=2).value = vendor.get('order_count', 0)
            on_time_pct = vendor.get('on_time_percentage', 0)
            ws.cell(row=row, column=3).value = on_time_pct / 100  # Format as percentage
            ws.cell(row=row, column=4).value = "Good" if on_time_pct >= 90 else "Fair" if on_time_pct >= 75 else "Poor"
            
            for col in range(1, 5):
                cell = ws.cell(row=row, column=col)
                cell.border = self.border
                if col == 3:
                    cell.number_format = '0%'
                    cell.alignment = Alignment(horizontal="right")
                elif col in [2]:
                    cell.alignment = Alignment(horizontal="right")
            
            row += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 14
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 20
    
    def _populate_low_stock_sheet(self, ws, items: List[Dict], start_row: int = 1):
        """Populate low stock items worksheet."""
        # Column headers
        headers = ["SKU", "Item Name", "Current Qty", "Minimum Qty", "Variance"]
        row = start_row
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
            cell.border = self.border
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Data rows
        row = start_row + 1
        for item in items:
            current = float(item.get('current', 0))
            minimum = float(item.get('minimum', 0))
            variance = current - minimum
            
            ws.cell(row=row, column=1).value = item.get('sku', '')
            ws.cell(row=row, column=2).value = item.get('name', '')
            ws.cell(row=row, column=3).value = current
            ws.cell(row=row, column=4).value = minimum
            ws.cell(row=row, column=5).value = variance
            
            for col in range(1, 6):
                cell = ws.cell(row=row, column=col)
                cell.border = self.border
                if col in [3, 4, 5]:
                    cell.alignment = Alignment(horizontal="right")
                if col == 5 and variance < 0:
                    cell.fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
            
            row += 1
        
        # Adjust column widths
        if start_row == 1:
            ws.column_dimensions['A'].width = 12
            ws.column_dimensions['B'].width = 25
            ws.column_dimensions['C'].width = 14
            ws.column_dimensions['D'].width = 14
            ws.column_dimensions['E'].width = 12
    
    def _create_analytics_overview_sheet(self, ws, analytics_data: Dict):
        """Create overview sheet for analytics report."""
        ws['A1'] = f"{self.company_name} Analytics Report"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:D1')
        
        ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws['A2'].font = Font(italic=True, size=10)
        
        row = 4
        
        # Order Metrics
        if 'order_metrics' in analytics_data:
            ws[f'A{row}'] = "Order Metrics"
            ws[f'A{row}'].font = Font(bold=True, size=11)
            row += 1
            
            order_stats = analytics_data['order_metrics']
            data = {
                "Total Orders": order_stats.get('total_orders', 0),
                "Pending Approvals": order_stats.get('pending_approvals', 0),
                "Avg Approval Time (days)": round(order_stats.get('average_approval_time_days', 0), 2),
                "Avg Dispatch Time (days)": round(order_stats.get('average_dispatch_time_days', 0), 2),
            }
            
            for key, value in data.items():
                ws[f'A{row}'] = key
                ws[f'B{row}'] = value
                row += 1
            
            row += 1
        
        # Inventory Health
        if 'inventory_health' in analytics_data:
            ws[f'A{row}'] = "Inventory Health"
            ws[f'A{row}'].font = Font(bold=True, size=11)
            row += 1
            
            inv_health = analytics_data['inventory_health']
            data = {
                "Total Items": inv_health.get('total_items', 0),
                "Low Stock Items": inv_health.get('low_stock_count', 0),
                "Total Quantity": round(float(inv_health.get('total_quantity', 0)), 0),
            }
            
            for key, value in data.items():
                ws[f'A{row}'] = key
                ws[f'B{row}'] = value
                row += 1
            
            row += 1
        
        # Email Stats
        if 'email_stats' in analytics_data:
            ws[f'A{row}'] = "Email Statistics"
            ws[f'A{row}'].font = Font(bold=True, size=11)
            row += 1
            
            email_stats = analytics_data['email_stats']
            data = {
                "Total Emails": email_stats.get('total_emails', 0),
                "Sent": email_stats.get('sent_count', 0),
                "Failed": email_stats.get('failed_count', 0),
            }
            
            for key, value in data.items():
                ws[f'A{row}'] = key
                ws[f'B{row}'] = value
                row += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 15


    def generate_custom_report(self, report_data: Dict) -> bytes:
        """Generate a custom-filtered report Excel workbook."""
        wb = Workbook()
        wb.remove(wb.active)

        period = report_data.get("period", {})

        # --- Orders sheet ---
        orders = report_data.get("orders", [])
        ws_orders = wb.active or wb.create_sheet("Orders")
        ws_orders.title = "Orders"
        headers = ["Order #", "Vendor", "Status", "Created", "Items Count"]
        for col, h in enumerate(headers, 1):
            cell = ws_orders.cell(row=1, column=col, value=h)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = self.border

        for row_idx, o in enumerate(orders, 2):
            ws_orders.cell(row=row_idx, column=1, value=o.get("order_number", "")).border = self.border
            ws_orders.cell(row=row_idx, column=2, value=o.get("vendor_name", "")).border = self.border
            ws_orders.cell(row=row_idx, column=3, value=o.get("status", "")).border = self.border
            ws_orders.cell(row=row_idx, column=4, value=str(o.get("created_at", ""))[:10]).border = self.border
            ws_orders.cell(row=row_idx, column=5, value=o.get("item_count", 0)).border = self.border

            # Items sub-table in notes column
            items = o.get("items", [])
            if items:
                notes = "; ".join(f"{i.get('sku','')} x{i.get('quantity_ordered',0)}" for i in items)
                ws_orders.cell(row=row_idx, column=6, value=notes).border = self.border

        ws_orders.column_dimensions['A'].width = 18
        ws_orders.column_dimensions['B'].width = 22
        ws_orders.column_dimensions['C'].width = 22
        ws_orders.column_dimensions['D'].width = 14
        ws_orders.column_dimensions['E'].width = 12
        ws_orders.column_dimensions['F'].width = 40

        # --- Inventory sheet ---
        inventory = report_data.get("inventory", [])
        ws_inv = wb.create_sheet("Inventory")
        inv_headers = ["SKU", "Name", "Current Qty", "Min Qty", "Reserved", "Category"]
        for col, h in enumerate(inv_headers, 1):
            cell = ws_inv.cell(row=1, column=col, value=h)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = self.border

        for row_idx, i in enumerate(inventory, 2):
            ws_inv.cell(row=row_idx, column=1, value=i.get("sku", "")).border = self.border
            ws_inv.cell(row=row_idx, column=2, value=i.get("name", "")).border = self.border
            ws_inv.cell(row=row_idx, column=3, value=i.get("current_quantity", 0)).border = self.border
            ws_inv.cell(row=row_idx, column=4, value=i.get("minimum_quantity", 0)).border = self.border
            ws_inv.cell(row=row_idx, column=5, value=i.get("reserved_quantity", 0)).border = self.border
            ws_inv.cell(row=row_idx, column=6, value=i.get("category", "")).border = self.border

        ws_inv.column_dimensions['A'].width = 14
        ws_inv.column_dimensions['B'].width = 22
        ws_inv.column_dimensions['C'].width = 12
        ws_inv.column_dimensions['D'].width = 10
        ws_inv.column_dimensions['E'].width = 12
        ws_inv.column_dimensions['F'].width = 18

        # --- Summary sheet ---
        ws_summary = wb.create_sheet("Summary")
        row = 1
        ws_summary.cell(row=row, column=1, value="Custom Report").font = Font(bold=True, size=14)
        row += 1
        ws_summary.cell(row=row, column=1, value=f"Period: {period.get('start','N/A')} to {period.get('end','N/A')}")
        row += 2
        ws_summary.cell(row=row, column=1, value="Metric").font = self.summary_font
        ws_summary.cell(row=row, column=2, value="Value").font = self.summary_font
        row += 1
        for key, val in [("Total Orders", len(orders)), ("Total Items", len(inventory))]:
            ws_summary.cell(row=row, column=1, value=key).border = self.border
            ws_summary.cell(row=row, column=2, value=val).border = self.border
            row += 1

        ws_summary.column_dimensions['A'].width = 25
        ws_summary.column_dimensions['B'].width = 15

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.getvalue()


def get_excel_report_generator() -> ExcelReportGenerator:
    """Factory function for Excel report generator."""
    return ExcelReportGenerator()
